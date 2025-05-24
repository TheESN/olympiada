from .models import Subdivision, School, Person, teacher
from .models import Application, Country, Student, Olympiada, Employee, Participant, Result, ROLES, sex
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill


class ProtocolGen():
    wb_font = Font(name='Times New Roman', size=11, color='00000000')
    wb_font_title = Font(name='Times New Roman', size=14, color='00000000', bold=True)
    wb_border = Border(left=Side(style='thin', color='00000000'), right=Side(style='thin', color='00000000'),
                       top=Side(style='thin', color='00000000'), bottom=Side(style='thin', color='00000000'))
    wb_alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    wb_purple_bg = PatternFill(fill_type='solid', start_color='ffccccff', end_color='ffccccff')

    @classmethod
    def generate(cls, olymp_id):
        wb = load_workbook(filename="./excel_templates/protocol_template.xlsx")
        ws = wb.active
        participant_list = Participant.objects.filter(application__olymp__id=olymp_id)
        cls.generate_protocol_file(ws, participant_list)

        return wb

    @classmethod
    def generate_protocol_file(cls, worksheet, participant_list):
        row = 7

        result_fields_added = False

        for i, participant in enumerate(participant_list):
            application = participant.application
            result_list = Result.objects.filter(participant=participant).order_by('task_number')

            cls.fill_application_info(row + i, application, worksheet)
            cls.fill_school_info(row + i, application.school, worksheet)
            cls.fill_subdivision_info(row + i, application.subdivision, worksheet)
            cls.fill_student_info(row + i, application.student, worksheet)
            cls.fill_employee_info(row + i, application.employee, worksheet)

            if not result_fields_added:
                for k, result in enumerate(result_list):
                    cls.fill_result_tasks(row - 1, result, worksheet, k)

                worksheet[chr(ord('A') + 15) + str(row - 2)] = "Заполняет жюри"
                worksheet[chr(ord('A') + 15) + str(row - 2)].fill = cls.wb_purple_bg
                worksheet[chr(ord('A') + 15) + str(row - 2)].font = cls.wb_font_title
                worksheet[chr(ord('A') + 15) + str(row - 2)].alignment = cls.wb_alignment

                worksheet.merge_cells(start_row=row - 2, start_column=16, end_row=row - 2, end_column=16 + result_list.count())
                worksheet[chr(ord('A') + 15) + str(row - 2)].border = cls.wb_border
                result_fields_added = True

            for k, result in enumerate(result_list):
                cls.fill_result_info(row + i, result, worksheet, k)

            cls.fill_result_sum(row+i, worksheet, result_list.count())
            
            for j in range(0, 17 + result_list.count()):
                worksheet[chr(ord('A') + j) + str(row + i)].font = cls.wb_font
                worksheet[chr(ord('A') + j) + str(row + i)].border = cls.wb_border
                worksheet[chr(ord('A') + j) + str(row + i)].alignment = cls.wb_alignment

                worksheet.row_dimensions[row+i].height = 36.8511

            worksheet[chr(ord('A')) + str(row + i)] = str(i+1)

    @classmethod
    def fill_application_info(cls, row, application, worksheet):
        participate = application.participate
        worksheet[chr(ord('A') + 12) + str(row)] = str(participate)

    @classmethod
    def fill_school_info(cls, row, school, worksheet):
        school_name = school.school_name
        worksheet[chr(ord('A') + 10) + str(row)] = str(school_name)

    @classmethod
    def fill_subdivision_info(cls, row, subdivision, worksheet):
        subdivision_name = subdivision.subdivision_name
        worksheet[chr(ord('A') + 9) + str(row)] = str(subdivision_name)

    @classmethod
    def fill_student_info(cls, row, student, worksheet):
        student_name_full = student.name.split()
        worksheet[chr(ord('A') + 2) + str(row)] = str(student_name_full[0])
        worksheet[chr(ord('A') + 3) + str(row)] = str(student_name_full[1])
        if len(student_name_full) == 3:
            worksheet[chr(ord('A') + 4) + str(row)] = str(student_name_full[2])

        student_sex = student.sex
        worksheet[chr(ord('A') + 5) + str(row)] = str(sex[student_sex][1][0])

        student_birthday = student.birthday
        worksheet[chr(ord('A') + 6) + str(row)] = str(student_birthday)

        student_country = student.country.country_name
        worksheet[chr(ord('A') + 7) + str(row)] = str(student_country)

        student_special_needs = student.special_needs
        worksheet[chr(ord('A') + 8) + str(row)] = "Да" if student_special_needs else "Нет"

        student_course = student.course_study
        worksheet[chr(ord('A') + 11) + str(row)] = str(student_course)

        student_phone = student.contact_phone
        worksheet[chr(ord('A') + 14) + str(row)] = str(student_phone)

    @classmethod
    def fill_employee_info(cls, row, employee, worksheet):
        employee_name = employee.name
        worksheet[chr(ord('A') + 13) + str(row)] = str(employee_name)

    @classmethod
    def fill_result_tasks(cls, row, result, worksheet, col):
        task_number = result.task_number
        worksheet.insert_cols(15 + col + 1)
        worksheet[chr(ord('A') + 14 + col + 1) + str(row)] = str(task_number)

        worksheet[chr(ord('A') + 14 + col + 1) + str(row)].fill = cls.wb_purple_bg
        worksheet[chr(ord('A') + 14 + col + 1) + str(row)].font = cls.wb_font_title
        worksheet[chr(ord('A') + 14 + col + 1) + str(row)].alignment = cls.wb_alignment
        worksheet[chr(ord('A') + 14 + col + 1) + str(row)].border = cls.wb_border

    @classmethod
    def fill_result_info(cls, row, result, worksheet, col):
        result_value = result.result_value
        worksheet[chr(ord('A') + 14 + col + 1) + str(row)] = float(result_value)

        # worksheet[chr(ord('A') + 14 + col + 1) + str(row)] += float(result_value)

    @classmethod
    def fill_result_sum(cls, row, worksheet, result_list_size):
        end_col = 14 + result_list_size
        start_col = 15
        worksheet[chr(ord('A') + end_col + 1) + str(row)] = f'=SUM({chr(ord('A') + start_col)}{row}:{chr(ord('A') + end_col)}{row})'