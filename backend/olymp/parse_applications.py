from openpyxl import load_workbook
from .factory import Factory

class ApplicationParser():
    @classmethod
    def parse_excel(cls, filename, olympiada, employee):
        wb = load_workbook(filename=filename)
        ws = wb.active
        return cls.load_applications(ws, olympiada, employee)

    @classmethod
    def load_applications(cls, worksheet, olympiada, employee):
        result = ""
        for row in worksheet.iter_rows(min_row=5, max_col=14):
            try:
                print("Calling Factory Function")
                result += Factory.create_application(row, olympiada, employee)
            except:
                result += "Не удалось добавить заявку " + row[1].value + " " + row[2].value + " " + row[3].value + "\n"
        return result