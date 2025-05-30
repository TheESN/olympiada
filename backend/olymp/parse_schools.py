from openpyxl import load_workbook
from .factory import Factory

class SchoolParser():
    @classmethod
    def parse_excel(cls, filename):
        wb = load_workbook(filename=filename)
        ws = wb.active
        return cls.load_schools(ws)

    @classmethod
    def load_schools(cls, worksheet):
        result = ""
        for row in worksheet.iter_rows(min_row=2):
            try:
                result += cls.load_school(row)
            except:
                result += "Не удалось добавить школу " + row[1].value + "\n"
        return result

    @classmethod
    def load_school(cls, row):
        return Factory.create_school(row[1].value, row[0].value)


