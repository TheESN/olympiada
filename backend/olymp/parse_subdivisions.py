from openpyxl import load_workbook
from .factory import Factory

class SubdivisionParser():
    @classmethod
    def parse_excel(cls, filename):
        wb = load_workbook(filename=filename)
        ws = wb.active
        return cls.load_subdivisions(ws)

    @classmethod
    def load_subdivisions(cls,worksheet):
        result = ""
        for row in worksheet.iter_rows(min_row=2):
                result += cls.load_subdivision(row)
        return result


    @classmethod
    def load_subdivision(cls, row):
        try:
            return Factory.create_subdivision(row[0].value)
        except:
            return "Не удалось создать муниципальное образование " + row[0].value